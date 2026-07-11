"""Build Git-friendly XLSForm CSV sources and the Kobo-ready XLSForm workbook."""

from __future__ import annotations

import csv
import re
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from project_config import DATASET_LABEL, REGIONS, REGION_LABELS, SURVEY_VERSION

ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "survey" / "source"
OUTPUT_PATH = ROOT / "survey" / "impact_survey_xlsform.xlsx"

SURVEY_HEADERS = [
    "type",
    "name",
    "label",
    "hint",
    "required",
    "relevant",
    "constraint",
    "constraint_message",
    "calculation",
    "appearance",
]

SURVEY_ROWS = [
    ["note", "synthetic_notice", DATASET_LABEL, "Synthetic dataset only; do not enter real personal information", "", "", "", "", "", ""],
    ["start", "start", "", "", "", "", "", "", "", ""],
    ["end", "end", "", "", "", "", "", "", "", ""],
    ["deviceid", "deviceid", "", "", "", "", "", "", "", ""],
    ["calculate", "synthetic_data", "", "", "", "", "", "", "'yes'", ""],
    ["calculate", "survey_version", "", "", "", "", "", "", f"'{SURVEY_VERSION}'", ""],
    ["select_one yes_no", "consent", "Do you agree to continue with this synthetic survey?", "Do not enter any real personal information.", "yes", "", "", "", "", ""],
    ["note", "no_consent_note", "Thank you. The survey ends here.", "", "", "${consent} = 'no'", "", "", "", ""],
    ["begin_group", "monitoring", "Monitoring questions", "", "", "${consent} = 'yes'", "", "", "", "field-list"],
    ["select_one survey_round", "survey_round", "Survey round", "", "yes", "", "", "", "", ""],
    ["date", "interview_date", "Interview date", "Use a date from 1 April to 30 September 2025.", "yes", "", ". >= date('2025-04-01') and . <= date('2025-09-30')", "Enter a date in the survey period.", "", ""],
    ["date", "program_start_date", "Activity start date", "", "yes", "", ". <= ${interview_date}", "Start date cannot be after the interview.", "", ""],
    ["select_one region", "region_code", "Fictional region", "", "yes", "", "", "", "", ""],
    ["select_one site", "site_code", "Fictional site", "", "yes", "", "", "", "", "minimal"],
    ["select_one enumerator", "enumerator_code", "Fictional enumerator code", "", "yes", "", "", "", "", ""],
    ["select_one age_group", "age_group", "Age group", "No exact birth date is collected.", "yes", "", "", "", "", ""],
    ["select_one gender", "gender", "Gender", "", "yes", "", "", "", "", ""],
    ["integer", "household_size", "Household size", "", "yes", "", ". >= 1 and . <= 20", "Enter a number from 1 to 20.", "", ""],
    ["select_one yes_no", "disability_access", "Was accessibility support needed?", "", "yes", "", "", "", "", ""],
    ["select_one yes_no", "participated_training", "Did the respondent participate in training?", "", "yes", "", "", "", "", ""],
    ["begin_group", "training", "Training follow-up", "", "", "${participated_training} = 'yes'", "", "", "", "field-list"],
    ["date", "training_date", "Training date", "", "yes", "", ". >= ${program_start_date} and . <= ${interview_date}", "Training must fall between the activity start and interview dates.", "", ""],
    ["select_one training_topic", "training_topic", "Main training topic", "", "yes", "", "", "", "", ""],
    ["integer", "satisfaction_score", "Satisfaction score", "1 = very dissatisfied; 5 = very satisfied", "yes", "", ". >= 1 and . <= 5", "Enter a score from 1 to 5.", "", "likert"],
    ["integer", "knowledge_before", "Knowledge before training", "1 = low; 5 = high", "yes", "", ". >= 1 and . <= 5", "Enter a score from 1 to 5.", "", "likert"],
    ["integer", "knowledge_after", "Knowledge after training", "1 = low; 5 = high", "yes", "", ". >= 1 and . <= 5", "Enter a score from 1 to 5.", "", "likert"],
    ["select_one yes_no", "used_skill", "Has a skill from the training been used?", "", "yes", "", "", "", "", ""],
    ["end_group", "training_end", "", "", "", "", "", "", "", ""],
    ["integer", "service_access", "Ease of accessing services", "1 = very difficult; 5 = very easy", "yes", "", ". >= 1 and . <= 5", "Enter a score from 1 to 5.", "", "likert"],
    ["select_one yes_no", "follow_up_requested", "Would the respondent like more general information?", "No contact details are collected.", "yes", "", "", "", "", ""],
    ["select_one preferred_channel", "preferred_channel", "Preferred non-personal information channel", "", "yes", "${follow_up_requested} = 'yes'", "", "", "", ""],
    ["text", "open_comment", "Optional comment", "Do not enter names or other identifying details.", "", "", "string-length(.) <= 300", "Use no more than 300 characters.", "", "multiline"],
    ["end_group", "monitoring_end", "", "", "", "", "", "", "", ""],
]

# Cascading site choices are kept as an explicit column because this is both a
# form-quality feature and a readable example of survey logic in Git.
SURVEY_HEADERS.append("choice_filter")
for survey_row in SURVEY_ROWS:
    survey_row.append("")
    if survey_row[1] == "site_code":
        survey_row[-1] = "region_code=${region_code}"

CHOICE_HEADERS = ["list_name", "name", "label"]
CHOICE_ROWS = [
    ["yes_no", "yes", "Yes"], ["yes_no", "no", "No"],
    ["survey_round", "ROUND_1", "Round 1"], ["survey_round", "ROUND_2", "Round 2"],
    *[["region", code, label] for code, label in REGION_LABELS.items()],
    *[["site", site, f"Site {site} (fictional)"] for sites in REGIONS.values() for site in sites],
    *[["enumerator", f"ENUM_{i:02d}", f"Enumerator {i:02d} (fictional)"] for i in range(1, 9)],
    ["age_group", "18_24", "18–24"], ["age_group", "25_34", "25–34"],
    ["age_group", "35_44", "35–44"], ["age_group", "45_54", "45–54"],
    ["age_group", "55_plus", "55 or older"],
    ["gender", "woman", "Woman"], ["gender", "man", "Man"],
    ["gender", "non_binary", "Non-binary"], ["gender", "prefer_not", "Prefer not to say"],
    ["training_topic", "planning", "Planning"],
    ["training_topic", "financial_skills", "Financial skills"],
    ["training_topic", "digital_skills", "Digital skills"],
    ["training_topic", "leadership", "Leadership"],
    ["preferred_channel", "community_meeting", "Community meeting"],
    ["preferred_channel", "printed_material", "Printed material"],
    ["preferred_channel", "radio", "Radio"],
]

CHOICE_HEADERS.append("region_code")
for choice_row in CHOICE_ROWS:
    choice_row.append("")
for choice_row in CHOICE_ROWS:
    if choice_row[0] == "site":
        choice_row[3] = next(
            region for region, sites in REGIONS.items() if choice_row[1] in sites
        )

SETTINGS_HEADERS = ["form_title", "form_id", "version", "instance_name", "default_language"]
SETTINGS_ROWS = [[
    "Impact Survey Data Quality Toolkit - Synthetic Training Follow-up Survey",
    "impact_survey_synthetic",
    SURVEY_VERSION,
    "concat('Synthetic survey - ', ${interview_date})",
    "English (en)",
]]


def write_csv(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(headers)
        writer.writerows(rows)


def add_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[str]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(sheet.columns, start=1):
        width = min(60, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[get_column_letter(index)].width = width


def normalize_xlsx_archive(path: Path) -> None:
    """Remove changing ZIP timestamps while preserving workbook content."""
    fixed_time = (1980, 1, 1, 0, 0, 0)
    members: list[tuple[str, int, int, bytes]] = []
    with zipfile.ZipFile(path, "r") as source:
        for name in sorted(source.namelist()):
            original = source.getinfo(name)
            content = source.read(name)
            if name == "docProps/core.xml":
                content = re.sub(
                    rb"(<dcterms:(created|modified)[^>]*>).*?(</dcterms:\2>)",
                    lambda match: (
                        match.group(1)
                        + b"2025-01-01T00:00:00Z"
                        + match.group(3)
                    ),
                    content,
                )
            members.append(
                (name, original.create_system, original.external_attr, content)
            )

    with zipfile.ZipFile(
        path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
    ) as target:
        for name, create_system, external_attr, content in members:
            info = zipfile.ZipInfo(name, date_time=fixed_time)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = create_system
            info.external_attr = external_attr
            target.writestr(info, content)


def build(
    output_path: Path = OUTPUT_PATH,
    source_dir: Path = SOURCE_DIR,
) -> Path:
    write_csv(source_dir / "survey.csv", SURVEY_HEADERS, SURVEY_ROWS)
    write_csv(source_dir / "choices.csv", CHOICE_HEADERS, CHOICE_ROWS)
    write_csv(source_dir / "settings.csv", SETTINGS_HEADERS, SETTINGS_ROWS)

    workbook = Workbook()
    workbook.remove(workbook.active)
    add_sheet(workbook, "survey", SURVEY_HEADERS, SURVEY_ROWS)
    add_sheet(workbook, "choices", CHOICE_HEADERS, CHOICE_ROWS)
    add_sheet(workbook, "settings", SETTINGS_HEADERS, SETTINGS_ROWS)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    normalize_xlsx_archive(output_path)
    return output_path


if __name__ == "__main__":
    path = build()
    print(f"Built XLSForm at {path}")
