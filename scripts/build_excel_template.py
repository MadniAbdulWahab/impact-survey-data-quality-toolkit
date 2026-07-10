"""Build the reproducible Excel monitoring template.

The workbook is a portfolio artefact, not a claim that Excel has been tested:
this script only writes deterministic ``.xlsx`` bytes. The interactive Power
Query, pivot, and VBA steps are performed and verified by a person, following
``excel/README.md``.

Design:

* Code lists, ranges, and the region-to-site map come from ``project_config``
  and the raw CSV, so the workbook cannot silently drift from the survey
  definition or the R pipeline.
* A self-contained ``Entry_Demo`` sheet carries a small labelled synthetic
  sample plus live formulas that mirror the seven R validation rules
  (``qc_missing_required`` ... ``qc_skip_logic``), data-validation dropdowns,
  and conditional formatting. It is openable and testable with no external
  wiring.
* Power Query (``excel/power_query/SurveyRaw.m``) is the production import path
  for the full 420-row extract; it feeds the pivots described in the README.
"""

from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from build_xlsform import normalize_xlsx_archive
from project_config import (
    AGE_GROUPS,
    DATASET_LABEL,
    ENUMERATORS,
    GENDERS,
    N_RESPONSES,
    PREFERRED_CHANNELS,
    REFERENCE_DATE,
    REGION_LABELS,
    REGIONS,
    SURVEY_END,
    SURVEY_ROUNDS,
    SURVEY_START,
    TRAINING_TOPICS,
    YES_NO,
)

ROOT = Path(__file__).resolve().parents[1]
RAW_CSV = ROOT / "data" / "raw" / "survey_responses_synthetic.csv"
OUTPUT_PATH = ROOT / "excel" / "impact_survey_monitoring_template.xlsx"

# Header styling reused from the XLSForm builder's visual language.
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
BANNER_FILL = PatternFill("solid", fgColor="FFF2CC")
FLAG_ROW_FILL = PatternFill("solid", fgColor="FCE4D6")
FLAG_TRUE_FILL = PatternFill("solid", fgColor="FFC7CE")

# The 27 raw columns, in file order, followed by the 9 QC columns produced by
# the R pipeline. The Excel demo reproduces the same 36-column schema.
QC_COLUMNS = [
    "qc_missing_required",
    "qc_duplicate_response_id",
    "qc_invalid_code",
    "qc_date_consistency",
    "qc_out_of_range",
    "qc_region_site_consistency",
    "qc_skip_logic",
    "qc_issue_count",
    "qc_any_issue",
]

# Curated sample: one or two rows per injected-issue category plus clean rows.
# Chosen from the committed raw extract so the demo always matches real data.
SAMPLE_SUBMISSION_IDS = [
    "SUB-0064",     # clean, participated in training
    "SUB-0002",     # clean, did not participate
    "SUB-0003",     # clean, accessibility support needed
    "SUB-0239",     # duplicate response id (original of SYN-0239)
    "SUB-DUP-001",  # duplicate response id (repeat of SYN-0239)
    "SUB-0116",     # missing_required: gender blank
    "SUB-0192",     # invalid_code: region CENTRAL
    "SUB-0009",     # date_consistency: training before programme start
    "SUB-0001",     # out_of_range: completion 180 minutes
    "SUB-0197",     # region_site_consistency: SOUTH region with site E01
    "SUB-0033",     # skip_logic: training fields set while participation is no
]

DATE_FIELDS = {"interview_date", "program_start_date", "training_date"}
INT_FIELDS = {
    "household_size",
    "satisfaction_score",
    "knowledge_before",
    "knowledge_after",
    "service_access",
}
FLOAT_FIELDS = {"completion_minutes"}

N_BLANK_ENTRY_ROWS = 20  # blank rows for a person to exercise the dropdowns


def load_sample_rows() -> tuple[list[str], list[dict[str, object]]]:
    """Read the raw CSV and return (headers, curated typed sample rows)."""
    with RAW_CSV.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        headers = list(reader.fieldnames or [])
        by_id = {row["submission_id"]: row for row in reader}

    rows: list[dict[str, object]] = []
    for submission_id in SAMPLE_SUBMISSION_IDS:
        raw = by_id[submission_id]
        rows.append({key: _coerce(key, value) for key, value in raw.items()})
    return headers, rows


def _coerce(field: str, value: str) -> object:
    """Convert raw CSV strings to real dates/numbers so formulas compare."""
    if value is None or value == "":
        return None
    if field in DATE_FIELDS:
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            return value  # keep deliberately malformed values visible
    if field in INT_FIELDS:
        try:
            return int(value)
        except ValueError:
            return value
    if field in FLOAT_FIELDS:
        try:
            return float(value)
        except ValueError:
            return value
    return value


def _style_header(sheet, row: int, columns: int) -> None:
    for index in range(1, columns + 1):
        cell = sheet.cell(row=row, column=index)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def _banner(sheet, text: str, span: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = sheet.cell(row=1, column=1, value=text)
    cell.font = Font(bold=True, color="7F6000")
    cell.fill = BANNER_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[1].height = 28


def _add_name(workbook: Workbook, name: str, ref: str) -> None:
    workbook.defined_names[name] = DefinedName(name=name, attr_text=ref)


def _auto_width(sheet, max_width: int = 42) -> None:
    for column_cells in sheet.columns:
        letter = None
        longest = 0
        for cell in column_cells:
            cell_letter = getattr(cell, "column_letter", None)
            if cell_letter:
                letter = cell_letter
            value = cell.value
            if value is not None:
                longest = max(longest, len(str(value)))
        if letter:
            sheet.column_dimensions[letter].width = min(max_width, max(12, longest + 2))


def build_start_here(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Start_Here")
    _banner(sheet, DATASET_LABEL, 2)
    lines = [
        ("What this is", "A reproducible monitoring workbook for the synthetic "
         "impact survey. Every value is fictional."),
        ("Self-contained demo", "The Entry_Demo sheet already works: it carries a "
         "labelled sample and live QC formulas, data validation, and conditional "
         "formatting. Nothing needs wiring to inspect it."),
        ("Full dataset", "To load all 420 rows, follow the numbered steps in "
         "excel/README.md: set the RawDataPath cell on Config, paste "
         "power_query/SurveyRaw.m into a Blank Query, then build the pivots."),
        ("Automation", "vba/RefreshAndExportQC.bas refreshes everything and writes "
         "a timestamped QC export into excel/exports/."),
        ("Parity", "The seven QC flags reproduce the R pipeline rules. See the "
         "QC_Rules sheet for the mapping."),
        ("Honesty", "This file is not marked 'tested' until a person completes the "
         "checkpoint in docs/verification/excel_test_log.md."),
    ]
    row = 3
    for heading, body in lines:
        head_cell = sheet.cell(row=row, column=1, value=heading)
        head_cell.font = Font(bold=True)
        head_cell.alignment = Alignment(vertical="top")
        body_cell = sheet.cell(row=row, column=2, value=body)
        body_cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row].height = 42
        row += 1
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 82
    sheet.sheet_view.showGridLines = False


def build_config(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Config")
    _banner(sheet, DATASET_LABEL, 3)
    sheet.cell(row=2, column=1, value="Parameter").font = Font(bold=True)
    sheet.cell(row=2, column=2, value="Value").font = Font(bold=True)
    sheet.cell(row=2, column=3, value="Notes").font = Font(bold=True)
    _style_header(sheet, 2, 3)

    survey_start = datetime.strptime(SURVEY_START, "%Y-%m-%d").date()
    survey_end = datetime.strptime(SURVEY_END, "%Y-%m-%d").date()
    reference_date = datetime.strptime(REFERENCE_DATE, "%Y-%m-%d").date()

    # (defined name, display label, value, note)
    params = [
        ("RawDataPath", "Raw data path",
         "data/raw/survey_responses_synthetic.csv",
         "Set to the full path on your machine before loading the query."),
        ("ExpectedRowCount", "Expected row count", N_RESPONSES,
         "Power Query and pivots should report this many rows."),
        ("SurveyStart", "Survey start", survey_start, "Earliest valid interview date."),
        ("SurveyEnd", "Survey end", survey_end, "Latest valid interview date."),
        ("ReferenceDate", "Reference date", reference_date, "Fixed analysis reference."),
        ("HouseholdMin", "Household minimum", 1, "Valid household size lower bound."),
        ("HouseholdMax", "Household maximum", 20, "Valid household size upper bound."),
        ("ScoreMin", "Score minimum", 1, "Valid 1-5 scale lower bound."),
        ("ScoreMax", "Score maximum", 5, "Valid 1-5 scale upper bound."),
        ("CompletionMin", "Completion minutes minimum", 3, "Valid interview duration floor."),
        ("CompletionMax", "Completion minutes maximum", 60, "Valid interview duration ceiling."),
    ]
    row = 3
    for name, label, value, note in params:
        sheet.cell(row=row, column=1, value=label)
        value_cell = sheet.cell(row=row, column=2, value=value)
        if isinstance(value, date):
            value_cell.number_format = "yyyy-mm-dd"
        note_cell = sheet.cell(row=row, column=3, value=note)
        note_cell.alignment = Alignment(wrap_text=True)
        _add_name(workbook, name, f"Config!$B${row}")
        row += 1

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 40
    sheet.column_dimensions["C"].width = 60
    sheet.freeze_panes = "A3"


def build_lookups(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Lookups")
    _banner(sheet, DATASET_LABEL + "  |  code lists generated from project_config", 8)

    sites = [site for sites in REGIONS.values() for site in sites]
    site_regions = [
        region for region, region_sites in REGIONS.items() for _ in region_sites
    ]

    # Each single-column list gets a header on row 2 and a defined name over its
    # body rows. Two-column blocks (labels, site->region map) sit side by side.
    def place(col: int, header: str, values: list[str], name: str | None) -> None:
        letter = get_column_letter(col)
        head = sheet.cell(row=2, column=col, value=header)
        head.font = Font(bold=True, color="FFFFFF")
        head.fill = HEADER_FILL
        for offset, value in enumerate(values):
            sheet.cell(row=3 + offset, column=col, value=value)
        if name:
            first, last = 3, 2 + len(values)
            _add_name(workbook, name, f"Lookups!${letter}${first}:${letter}${last}")

    place(1, "Region", list(REGION_LABELS.keys()), "ValidRegions")
    place(2, "Region label", list(REGION_LABELS.values()), None)
    place(3, "Site", sites, "ValidSites")
    place(4, "Site region", site_regions, None)
    place(5, "Round", list(SURVEY_ROUNDS), "ValidRounds")
    place(6, "Yes/No", list(YES_NO), "ValidYesNo")
    place(7, "Age group", list(AGE_GROUPS), "ValidAgeGroups")
    place(8, "Gender", list(GENDERS), "ValidGenders")
    place(9, "Training topic", list(TRAINING_TOPICS), "ValidTopics")
    place(10, "Preferred channel", list(PREFERRED_CHANNELS), "ValidChannels")
    place(11, "Enumerator", list(ENUMERATORS), "ValidEnumerators")

    # SiteList / SiteRegionList are the two columns of the site->region map used
    # by the region/site consistency formula.
    _add_name(workbook, "SiteList", f"Lookups!$C$3:$C${2 + len(sites)}")
    _add_name(workbook, "SiteRegionList", f"Lookups!$D$3:$D${2 + len(sites)}")

    _auto_width(sheet)
    sheet.freeze_panes = "A3"


def build_qc_rules(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("QC_Rules")
    _banner(sheet, DATASET_LABEL + "  |  Excel formulas mirror these R pipeline rules", 4)
    headers = ["QC column", "Fields", "Rule", "Consent-gated"]
    for index, text in enumerate(headers, start=1):
        sheet.cell(row=2, column=index, value=text)
    _style_header(sheet, 2, len(headers))

    rules = [
        ("qc_missing_required",
         "all required fields",
         "A required field is blank given the skip conditions.", "Yes"),
        ("qc_duplicate_response_id",
         "response_id",
         "response_id appears on more than one record.", "No"),
        ("qc_invalid_code",
         "coded fields",
         "A non-blank code is not in its allowed list.", "No"),
        ("qc_date_consistency",
         "interview / start / training dates",
         "Interview outside the survey window, start after interview, or "
         "training outside the start-to-interview span.", "Yes"),
        ("qc_out_of_range",
         "household, 1-5 scales, completion minutes",
         "A numeric value falls outside its documented bounds.", "Yes"),
        ("qc_region_site_consistency",
         "region_code, site_code",
         "The site does not belong to the selected region.", "Yes"),
        ("qc_skip_logic",
         "training / follow-up fields",
         "Fields are populated that the skip logic should leave blank.", "Yes"),
    ]
    row = 3
    for qc_column, fields, rule, gated in rules:
        sheet.cell(row=row, column=1, value=qc_column).font = Font(name="Consolas")
        sheet.cell(row=row, column=2, value=fields)
        rule_cell = sheet.cell(row=row, column=3, value=rule)
        rule_cell.alignment = Alignment(wrap_text=True)
        sheet.cell(row=row, column=4, value=gated)
        row += 1
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 62
    sheet.column_dimensions["D"].width = 15
    sheet.freeze_panes = "A3"


# Column letters for the Entry_Demo schema, keyed by field name.
def _column_letters(headers: list[str]) -> dict[str, str]:
    return {name: get_column_letter(index) for index, name in enumerate(headers, start=1)}


def _qc_formulas(row: int) -> dict[str, str]:
    """Return the nine QC formulas for a given data row (1-indexed)."""
    guard = f"COUNTA($A{row}:$AA{row})=0"

    invalid_parts = [
        f'$C{row}<>"yes"',
        f'AND($E{row}<>"",ISNA(MATCH($E{row},ValidRounds,0)))',
        f'AND($H{row}<>"",ISNA(MATCH($H{row},ValidRegions,0)))',
        f'AND($I{row}<>"",ISNA(MATCH($I{row},ValidSites,0)))',
        f'AND($J{row}<>"",ISNA(MATCH($J{row},ValidEnumerators,0)))',
        f'AND($K{row}<>"",ISNA(MATCH($K{row},ValidYesNo,0)))',
        f'AND($L{row}<>"",ISNA(MATCH($L{row},ValidAgeGroups,0)))',
        f'AND($M{row}<>"",ISNA(MATCH($M{row},ValidGenders,0)))',
        f'AND($O{row}<>"",ISNA(MATCH($O{row},ValidYesNo,0)))',
        f'AND($P{row}<>"",ISNA(MATCH($P{row},ValidYesNo,0)))',
        f'AND($R{row}<>"",ISNA(MATCH($R{row},ValidTopics,0)))',
        f'AND($V{row}<>"",ISNA(MATCH($V{row},ValidYesNo,0)))',
        f'AND($X{row}<>"",ISNA(MATCH($X{row},ValidYesNo,0)))',
        f'AND($Y{row}<>"",ISNA(MATCH($Y{row},ValidChannels,0)))',
    ]

    core_missing = (
        f"(COUNTBLANK($E{row}:$J{row})+COUNTBLANK($L{row}:$P{row})"
        f"+COUNTBLANK($W{row}:$X{row})+COUNTBLANK($AA{row}))>0"
    )
    training_missing = (
        f'AND($P{row}="yes",(COUNTBLANK($Q{row}:$R{row})'
        f"+COUNTBLANK($S{row}:$V{row}))>0)"
    )
    follow_up_missing = f'AND($X{row}="yes",$Y{row}="")'

    date_rule = (
        f'AND($K{row}="yes",OR('
        f'AND($F{row}<>"",OR($F{row}<SurveyStart,$F{row}>SurveyEnd)),'
        f'AND($G{row}<>"",$F{row}<>"",$G{row}>$F{row}),'
        f'AND($P{row}="yes",$Q{row}<>"",$G{row}<>"",$Q{row}<$G{row}),'
        f'AND($P{row}="yes",$Q{row}<>"",$F{row}<>"",$Q{row}>$F{row})))'
    )
    range_rule = (
        f'AND($K{row}="yes",OR('
        f'AND($N{row}<>"",OR($N{row}<HouseholdMin,$N{row}>HouseholdMax)),'
        f'AND($S{row}<>"",OR($S{row}<ScoreMin,$S{row}>ScoreMax)),'
        f'AND($T{row}<>"",OR($T{row}<ScoreMin,$T{row}>ScoreMax)),'
        f'AND($U{row}<>"",OR($U{row}<ScoreMin,$U{row}>ScoreMax)),'
        f'AND($W{row}<>"",OR($W{row}<ScoreMin,$W{row}>ScoreMax)),'
        f'AND($AA{row}<>"",OR($AA{row}<CompletionMin,$AA{row}>CompletionMax))))'
    )
    region_site_rule = (
        f'AND($K{row}="yes",$H{row}<>"",$I{row}<>"",'
        f'IFERROR(INDEX(SiteRegionList,MATCH($I{row},SiteList,0)),"")<>$H{row})'
    )
    skip_rule = (
        f'AND($K{row}="yes",OR('
        f'AND($P{row}="no",OR($Q{row}<>"",$R{row}<>"",$S{row}<>"",'
        f'$T{row}<>"",$U{row}<>"",$V{row}<>"")),'
        f'AND($X{row}="no",$Y{row}<>"")))'
    )

    return {
        "qc_missing_required":
            f'=IF({guard},"",AND($K{row}="yes",'
            f"OR({core_missing},{training_missing},{follow_up_missing})))",
        "qc_duplicate_response_id":
            f'=IF({guard},"",AND($B{row}<>"",COUNTIF($B$3:$B${LAST_DATA_ROW},$B{row})>1))',
        "qc_invalid_code":
            f'=IF({guard},"",OR({",".join(invalid_parts)}))',
        "qc_date_consistency": f'=IF({guard},"",{date_rule})',
        "qc_out_of_range": f'=IF({guard},"",{range_rule})',
        "qc_region_site_consistency": f'=IF({guard},"",{region_site_rule})',
        "qc_skip_logic": f'=IF({guard},"",{skip_rule})',
        "qc_issue_count": f'=IF({guard},"",COUNTIF($AB{row}:$AH{row},TRUE))',
        "qc_any_issue": f'=IF({guard},"",$AI{row}>0)',
    }


# Filled in by build() once the sample size is known.
LAST_DATA_ROW = 33


def build_entry_demo(workbook: Workbook, headers: list[str], rows: list[dict[str, object]]) -> None:
    global LAST_DATA_ROW
    sheet = workbook.create_sheet("Entry_Demo")
    all_headers = headers + QC_COLUMNS
    total_columns = len(all_headers)
    _banner(
        sheet,
        DATASET_LABEL + "  |  sample rows 3-{n} are known issues; rows below are yours to try"
        .format(n=2 + len(rows)),
        total_columns,
    )

    # Header row.
    for index, name in enumerate(all_headers, start=1):
        sheet.cell(row=2, column=index, value=name)
    _style_header(sheet, 2, total_columns)

    first_row = 3
    last_row = first_row + len(rows) + N_BLANK_ENTRY_ROWS - 1
    LAST_DATA_ROW = last_row

    # Sample data.
    for offset, record in enumerate(rows):
        excel_row = first_row + offset
        for col_index, name in enumerate(headers, start=1):
            value = record.get(name)
            cell = sheet.cell(row=excel_row, column=col_index, value=value)
            if name in DATE_FIELDS and isinstance(value, date):
                cell.number_format = "yyyy-mm-dd"

    # QC formulas for every data row (sample + blank entry rows).
    qc_start = len(headers) + 1
    for excel_row in range(first_row, last_row + 1):
        formulas = _qc_formulas(excel_row)
        for offset, qc_name in enumerate(QC_COLUMNS):
            sheet.cell(row=excel_row, column=qc_start + offset, value=formulas[qc_name])

    _add_data_validation(sheet, headers, first_row, last_row)
    _add_conditional_formatting(sheet, first_row, last_row)

    sheet.freeze_panes = "B3"
    _auto_width(sheet, max_width=22)


def _add_data_validation(sheet, headers: list[str], first_row: int, last_row: int) -> None:
    letters = _column_letters(headers)
    list_rules = {
        "survey_round": "ValidRounds",
        "region_code": "ValidRegions",
        "site_code": "ValidSites",
        "enumerator_code": "ValidEnumerators",
        "consent": "ValidYesNo",
        "age_group": "ValidAgeGroups",
        "gender": "ValidGenders",
        "disability_access": "ValidYesNo",
        "participated_training": "ValidYesNo",
        "training_topic": "ValidTopics",
        "used_skill": "ValidYesNo",
        "follow_up_requested": "ValidYesNo",
        "preferred_channel": "ValidChannels",
    }
    for field, name in list_rules.items():
        dv = DataValidation(type="list", formula1=name, allow_blank=True)
        dv.error = f"Choose a value from the {field} list."
        dv.errorTitle = "Invalid code"
        dv.prompt = f"Select a valid {field}."
        sheet.add_data_validation(dv)
        dv.add(f"{letters[field]}{first_row}:{letters[field]}{last_row}")

    whole_rules = {
        "household_size": ("HouseholdMin", "HouseholdMax"),
        "satisfaction_score": ("ScoreMin", "ScoreMax"),
        "knowledge_before": ("ScoreMin", "ScoreMax"),
        "knowledge_after": ("ScoreMin", "ScoreMax"),
        "service_access": ("ScoreMin", "ScoreMax"),
    }
    for field, (low, high) in whole_rules.items():
        dv = DataValidation(
            type="whole", operator="between", formula1=low, formula2=high, allow_blank=True
        )
        dv.error = f"{field} must be between the configured bounds."
        dv.errorTitle = "Out of range"
        sheet.add_data_validation(dv)
        dv.add(f"{letters[field]}{first_row}:{letters[field]}{last_row}")

    completion = DataValidation(
        type="decimal", operator="between", formula1="CompletionMin",
        formula2="CompletionMax", allow_blank=True,
    )
    completion.error = "completion_minutes must be within the configured bounds."
    completion.errorTitle = "Out of range"
    sheet.add_data_validation(completion)
    completion.add(f"{letters['completion_minutes']}{first_row}:{letters['completion_minutes']}{last_row}")

    interview = DataValidation(
        type="date", operator="between", formula1="SurveyStart",
        formula2="SurveyEnd", allow_blank=True,
    )
    interview.error = "interview_date must fall inside the survey window."
    interview.errorTitle = "Date outside window"
    sheet.add_data_validation(interview)
    interview.add(f"{letters['interview_date']}{first_row}:{letters['interview_date']}{last_row}")


def _add_conditional_formatting(sheet, first_row: int, last_row: int) -> None:
    # Highlight any record that carries at least one QC flag.
    record_range = f"A{first_row}:AA{last_row}"
    sheet.conditional_formatting.add(
        record_range,
        FormulaRule(formula=[f"$AJ{first_row}=TRUE"], fill=FLAG_ROW_FILL),
    )
    # Highlight the individual flag cells that are TRUE.
    flag_range = f"AB{first_row}:AH{last_row}"
    sheet.conditional_formatting.add(
        flag_range,
        FormulaRule(formula=[f"AB{first_row}=TRUE"], fill=FLAG_TRUE_FILL),
    )


def build_monitoring_summary(workbook: Workbook, first_row: int, last_row: int) -> None:
    sheet = workbook.create_sheet("MonitoringSummary")
    _banner(
        sheet,
        DATASET_LABEL + "  |  reflects the Entry_Demo sample only; full 420-row "
        "monitoring uses the Power Query pivots (see excel/README.md)",
        6,
    )
    headers = [
        "Region", "Round 1", "Round 2", "Total responses",
        "Participated", "Mean satisfaction",
    ]
    for index, text in enumerate(headers, start=1):
        sheet.cell(row=2, column=index, value=text)
    _style_header(sheet, 2, len(headers))

    region_col = "H"
    round_col = "E"
    part_col = "P"
    sat_col = "S"

    def col(letter: str) -> str:
        return f"Entry_Demo!${letter}${first_row}:${letter}${last_row}"

    row = 3
    for region in REGION_LABELS:
        sheet.cell(row=row, column=1, value=region)
        sheet.cell(
            row=row, column=2,
            value=f'=COUNTIFS({col(region_col)},$A{row},{col(round_col)},"ROUND_1")',
        )
        sheet.cell(
            row=row, column=3,
            value=f'=COUNTIFS({col(region_col)},$A{row},{col(round_col)},"ROUND_2")',
        )
        sheet.cell(row=row, column=4, value=f"=B{row}+C{row}")
        sheet.cell(
            row=row, column=5,
            value=f'=COUNTIFS({col(region_col)},$A{row},{col(part_col)},"yes")',
        )
        sheet.cell(
            row=row, column=6,
            value=f'=IFERROR(AVERAGEIFS({col(sat_col)},{col(region_col)},$A{row},'
                  f'{col(sat_col)},">0"),"n/a")',
        )
        row += 1

    # Column totals for the demo sample.
    sheet.cell(row=row, column=1, value="All regions").font = Font(bold=True)
    for column in range(2, 6):
        letter = get_column_letter(column)
        sheet.cell(row=row, column=column, value=f"=SUM({letter}3:{letter}{row - 1})").font = Font(bold=True)

    for column in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 18
    sheet.freeze_panes = "A3"


def build(output_path: Path = OUTPUT_PATH) -> Path:
    headers, rows = load_sample_rows()

    workbook = Workbook()
    workbook.remove(workbook.active)
    build_start_here(workbook)
    build_config(workbook)
    build_lookups(workbook)
    build_qc_rules(workbook)
    build_entry_demo(workbook, headers, rows)
    build_monitoring_summary(workbook, first_row=3, last_row=LAST_DATA_ROW)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    normalize_xlsx_archive(output_path)
    return output_path


if __name__ == "__main__":
    path = build()
    print(f"Built Excel monitoring template at {path}")
