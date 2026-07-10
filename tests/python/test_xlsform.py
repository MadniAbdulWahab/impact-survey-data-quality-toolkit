from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from build_xlsform import build

ROOT = Path(__file__).resolve().parents[2]


def sheet_records(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    headers = list(rows[0])
    return [dict(zip(headers, row)) for row in rows[1:]]


def test_xlsform_has_required_sheets_and_headers():
    path = build()
    workbook = load_workbook(path, data_only=False)
    assert {"survey", "choices", "settings"}.issubset(workbook.sheetnames)
    assert {"type", "name", "label"}.issubset(
        {cell.value for cell in workbook["survey"][1]}
    )
    assert {"list_name", "name", "label"}.issubset(
        {cell.value for cell in workbook["choices"][1]}
    )
    assert {"form_title", "form_id", "version"}.issubset(
        {cell.value for cell in workbook["settings"][1]}
    )


def test_select_lists_exist_and_groups_are_balanced():
    path = ROOT / "survey" / "impact_survey_xlsform.xlsx"
    workbook = load_workbook(path, data_only=False)
    survey = sheet_records(workbook["survey"])
    choices = sheet_records(workbook["choices"])

    choice_lists = {row["list_name"] for row in choices}
    referenced_lists = {
        row["type"].split(maxsplit=1)[1]
        for row in survey
        if isinstance(row["type"], str) and row["type"].startswith("select_one ")
    }
    assert referenced_lists <= choice_lists

    group_counts = Counter(
        "begin" if row["type"] == "begin_group" else "end"
        for row in survey
        if row["type"] in {"begin_group", "end_group"}
    )
    assert group_counts["begin"] == group_counts["end"]


def test_form_has_privacy_notice_and_no_direct_identifier_questions():
    workbook = load_workbook(ROOT / "survey" / "impact_survey_xlsform.xlsx", data_only=False)
    survey = sheet_records(workbook["survey"])
    names = {row["name"] for row in survey}
    forbidden = {"name", "full_name", "phone", "email", "address", "gps", "date_of_birth"}
    assert forbidden.isdisjoint(names)
    notice = next(row for row in survey if row["name"] == "synthetic_notice")
    assert "SYNTHETIC DATA" in notice["label"]


def test_site_choices_use_region_filter():
    workbook = load_workbook(ROOT / "survey" / "impact_survey_xlsform.xlsx", data_only=False)
    survey = sheet_records(workbook["survey"])
    choices = sheet_records(workbook["choices"])
    site_question = next(row for row in survey if row["name"] == "site_code")
    assert site_question["choice_filter"] == "region_code=${region_code}"
    assert all(row["region_code"] for row in choices if row["list_name"] == "site")

