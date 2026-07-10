from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from build_excel_template import (
    QC_COLUMNS,
    SAMPLE_SUBMISSION_IDS,
    build,
    load_sample_rows,
)
from project_config import (
    AGE_GROUPS,
    GENDERS,
    PREFERRED_CHANNELS,
    REGIONS,
    TRAINING_TOPICS,
)

ROOT = Path(__file__).resolve().parents[2]
VERIFIED_WORKBOOK = ROOT / "excel" / "impact_survey_monitoring.xlsm"

REQUIRED_SHEETS = {
    "Start_Here",
    "Config",
    "Lookups",
    "QC_Rules",
    "Entry_Demo",
    "MonitoringSummary",
}

REQUIRED_NAMES = {
    "RawDataPath",
    "ExpectedRowCount",
    "SurveyStart",
    "SurveyEnd",
    "HouseholdMin",
    "HouseholdMax",
    "ScoreMin",
    "ScoreMax",
    "CompletionMin",
    "CompletionMax",
    "ValidRegions",
    "ValidSites",
    "ValidRounds",
    "ValidYesNo",
    "ValidAgeGroups",
    "ValidGenders",
    "ValidTopics",
    "ValidChannels",
    "ValidEnumerators",
    "SiteList",
    "SiteRegionList",
}


def test_template_has_required_sheets_and_names(tmp_path):
    path = build(tmp_path / "template.xlsx")
    workbook = load_workbook(path, data_only=False)
    assert REQUIRED_SHEETS.issubset(workbook.sheetnames)
    assert REQUIRED_NAMES.issubset(set(workbook.defined_names))


def test_entry_demo_schema_matches_pipeline(tmp_path):
    path = build(tmp_path / "template.xlsx")
    workbook = load_workbook(path, data_only=False)
    headers = [cell.value for cell in workbook["Entry_Demo"][2]]
    # 27 raw columns then the 9 QC columns, in the same order as the R output.
    assert headers[-len(QC_COLUMNS):] == QC_COLUMNS
    assert headers[0] == "submission_id"
    assert headers[26] == "completion_minutes"


def test_lookups_match_project_config(tmp_path):
    path = build(tmp_path / "template.xlsx")
    workbook = load_workbook(path, data_only=False)
    lookups = workbook["Lookups"]

    def column_values(letter: str) -> list[str]:
        values = []
        for row in range(3, lookups.max_row + 1):
            value = lookups[f"{letter}{row}"].value
            if value is not None:
                values.append(value)
        return values

    expected_sites = [site for sites in REGIONS.values() for site in sites]
    assert column_values("A") == list(REGIONS.keys())   # regions
    assert column_values("C") == expected_sites         # sites
    assert column_values("E") == ["ROUND_1", "ROUND_2"]  # rounds
    assert column_values("F") == ["yes", "no"]           # yes/no
    assert column_values("G") == list(AGE_GROUPS)        # age groups
    assert column_values("H") == list(GENDERS)           # genders
    assert column_values("I") == list(TRAINING_TOPICS)   # training topics
    assert column_values("J") == list(PREFERRED_CHANNELS)  # channels


def test_site_region_map_is_consistent(tmp_path):
    path = build(tmp_path / "template.xlsx")
    workbook = load_workbook(path, data_only=False)
    lookups = workbook["Lookups"]
    site_to_region = {}
    for row in range(3, lookups.max_row + 1):
        site = lookups[f"C{row}"].value
        region = lookups[f"D{row}"].value
        if site is not None:
            site_to_region[site] = region
    for region, sites in REGIONS.items():
        for site in sites:
            assert site_to_region[site] == region


def test_sample_covers_every_injected_issue_category():
    truth_path = ROOT / "data" / "raw" / "injected_issue_truth_synthetic.csv"
    lines = truth_path.read_text(encoding="utf-8").splitlines()[1:]
    sample = set(SAMPLE_SUBMISSION_IDS)
    covered = set()
    for line in lines:
        issue_type, submission_id, *_ = line.split(",")
        if submission_id in sample:
            covered.add(issue_type)
    expected = {
        "missing_required",
        "duplicate_response_id",
        "invalid_code",
        "inconsistent_date",
        "out_of_range",
        "consistency_violation",
        "skip_logic_violation",
    }
    assert expected.issubset(covered)


def test_entry_demo_qc_columns_hold_formulas(tmp_path):
    path = build(tmp_path / "template.xlsx")
    workbook = load_workbook(path, data_only=False)
    demo = workbook["Entry_Demo"]
    # First sample row lives on row 3; every QC cell should be a formula.
    for offset in range(len(QC_COLUMNS)):
        cell = demo.cell(row=3, column=28 + offset)
        assert isinstance(cell.value, str) and cell.value.startswith("=")


def test_build_is_byte_reproducible(tmp_path):
    first = build(tmp_path / "first.xlsx")
    second = build(tmp_path / "second.xlsx")
    assert first.read_bytes() == second.read_bytes()


def test_verified_macro_workbook_contains_reviewed_excel_features():
    assert VERIFIED_WORKBOOK.exists()
    with ZipFile(VERIFIED_WORKBOOK) as archive:
        names = archive.namelist()
        assert "xl/vbaProject.bin" in names
        assert "xl/connections.xml" in names
        assert "xl/queryTables/queryTable1.xml" in names
        assert "xl/pivotTables/pivotTable1.xml" in names
        assert "xl/pivotTables/pivotTable2.xml" in names
        assert "xl/pivotCache/pivotCacheDefinition1.xml" in names
        assert "xl/pivotCache/pivotCacheDefinition2.xml" in names

        vba = archive.read("xl/vbaProject.bin")
        assert b"modQCExport" in vba
        assert b"RefreshAndExportQC" in vba
        connection = archive.read("xl/connections.xml")
        assert b"Location=SurveyRaw" in connection
        for cache_number in (1, 2):
            cache = archive.read(
                f"xl/pivotCache/pivotCacheDefinition{cache_number}.xml"
            )
            assert b'recordCount="420"' in cache

    workbook = load_workbook(
        VERIFIED_WORKBOOK,
        keep_vba=True,
        data_only=True,
    )
    assert {
        "SurveyData",
        "Pivot_Region_Round",
        "Pivot_Satisfaction",
        "QC_Export",
    }.issubset(workbook.sheetnames)
    assert workbook["SurveyData"].max_row == 421
    assert "SurveyTbl" in workbook["SurveyData"].tables
    assert len(workbook["Pivot_Region_Round"]._pivots) == 1
    assert len(workbook["Pivot_Satisfaction"]._pivots) == 1
    assert workbook["Pivot_Region_Round"]["E10"].value == 420
    assert workbook["QC_Export"]["B10"].value == 8
    assert str(workbook["Config"]["B3"].value).startswith("SET_FULL_PATH_TO")
    workbook.close()


def test_verified_macro_workbook_has_no_local_machine_metadata():
    forbidden = (
        b"C:\\Users",
        "C:\\Users".encode("utf-16le"),
    )
    with ZipFile(VERIFIED_WORKBOOK) as archive:
        for name in archive.namelist():
            content = archive.read(name)
            assert not any(value in content for value in forbidden), name
        core_properties = archive.read("docProps/core.xml")
        assert b"<dc:creator></dc:creator>" in core_properties
        assert b"<cp:lastModifiedBy></cp:lastModifiedBy>" in core_properties
